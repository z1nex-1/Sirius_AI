import openai
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import seaborn as sns
import numpy as np
import re
from wordcloud import WordCloud, STOPWORDS
from matplotlib.colors import LinearSegmentedColormap
from collections import Counter
import os

openai.api_base = "http://localhost:1234/v1"
openai.api_key = "not-needed"
NUM_REVIEWS = 50
NUM_WORDS = 40
WORD_CLOUD_WIDTH = 1000
WORD_CLOUD_HEIGHT = 700
FIGSIZE = (16, 12)
HEIGHT_RATIOS = [1, 6]
FONTSIZE = 12
hspace = 0.05
print("Ключ API OpenAI загружен успешно")

df = pd.read_csv("samples.csv", delimiter=';', names=['id', 'username', 'review'], encoding='utf-8').head(NUM_REVIEWS)
print("CSV файл загружен успешно")
print(f"Количество отзывов: {len(df)}")

wb = openpyxl.Workbook()
sheet = wb.active


def analyze_problems(review):
    messages = [
        {"role": "system", "content": "Вы помощник в определении проблемы, с которой столкунлся пользователем в отзывах. Ваш ответ должен содержать только проблему. Ответ должен быть только на русском языке и должен быть длиной максимум 20 слов. Если не возможно определить, что не понравилось пользователю, выведи -"},        {"role": "user", "content": f"В очень краткой форме определи что не понравилось пользователю, и сделай вывод, что можно исправить. Если не возможно определить что не понравилось ползователю, выведи -. Основывайся вот на этом отзыве: {review}"}
    ]

    try:
        completion = openai.ChatCompletion.create(
            model="local-model",
            messages=messages,
            temperature=0,
        )

        response = completion.choices[0].message.content.strip()

        sheet.cell(row=sheet.max_row + 1, column=1, value=response)
        wb.save('weutput.xlsx')
    except Exception as e:
        print(f"Ошибка: {e}")
        return np.nan


df['problems'] = df['review'].apply(analyze_problems)
